import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\28129\Desktop\共享1-数据收集与核对\核验结果\最终企业样本名单.xlsx')
ws = wb.active

# 3 brown -> green
changes = {
    '600126': ('绿色', '宁波基地环保搬迁升级,超低排放改造,转为绿色组'),
    '600531': ('绿色', '有色冶炼清洁生产标杆,省级绿色工厂,转为绿色组'),
    '600409': ('绿色', '纯碱行业龙头,清洁生产认证,转为绿色组'),
}

fixed = 0
for r in range(5, ws.max_row+1):
    code = str(ws.cell(row=r, column=4).value or '').strip()
    if code in changes:
        new_group, reason = changes[code]
        ws.cell(row=r, column=3).value = new_group
        ws.cell(row=r, column=9).value = reason
        print(f'Row {r}: {code} brown -> {new_group}')
        fixed += 1

# Update title
ws.cell(row=2, column=1).value = '建模样本50家(绿20+争议20+棕10) | 行业标杆15家 | 独立测试集10家 | 更新日期:2026-06-03'

wb.save(r'C:\Users\28129\Desktop\共享1-数据收集与核对\核验结果\最终企业样本名单.xlsx')
print(f'Fixed {fixed} companies. Green=20 Brown=10 now.')
